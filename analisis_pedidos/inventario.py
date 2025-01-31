import aiohttp
import asyncio
import pandas as pd
import os
from openpyxl import load_workbook
import logging

# Configuración de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Configuración de la API
url = "http://201.234.74.137:82/v3/ejecutarconsultaestandar"
headers = {
    'conniKey': 'Connikey-parasolestropicalesdamis-VJDSNLE1',
    'conniToken': 'VJDSNLE1UDVAOFG4SZNFMU40SDJYOFG4SZNAOESZWTHPNFO4VTDZOA'
}
params = {
    'idCompania': 6631,
    'descripcion': 'API_v2_Inventarios_InvFecha',
    'parametros': "f150_id=''MP002'' and f400_cant_existencia_1 > 0"
}
num_pag = 1  # Número de página
tam_pag = 50  # Tamaño de página
max_concurrent_tasks = 1  # Máximo de tareas concurrentes

ruta_archivo = r"C:\Users\JORGE CONTRERAS\OneDrive - 900208659-2 DAMIS SAS\Escritorio\PLANEACION\consumos\INDICADORES PLANEACION\analisis_pedidos.xlsx"
hoja_objetivo = "inventarios"


async def fetch_page(session, num_pag):
    """
    Función para obtener una página de datos de la API.
    """
    params_with_paging = {**params, 'paginacion': f'numPag={num_pag}|tamPag={tam_pag}'}
    try:
        async with session.get(url, headers=headers, params=params_with_paging) as response:
            logging.info(f"Consultando página {num_pag} con parámetros: {params_with_paging}")
            if response.status != 200:
                logging.error(f"Error al obtener la página {num_pag}: {response.status} - {await response.text()}")
                return []
            data = await response.json()
            if not isinstance(data, dict) or 'detalle' not in data or 'Table' not in data['detalle']:
                logging.warning(f"Respuesta inesperada en la página {num_pag}: {data}")
                return []
            return data['detalle']['Table']
    except Exception as e:
        logging.error(f"Excepción al obtener la página {num_pag}: {e}")
        return []


async def fetch_all_data():
    """
    Función principal para obtener todos los datos de forma concurrente.
    """
    base_inventario = []
    async with aiohttp.ClientSession() as session:
        page = 1
        while True:
            logging.info(f"Consultando páginas {page} a {page + max_concurrent_tasks - 1}")
            tasks = [
                fetch_page(session, page + i)
                for i in range(max_concurrent_tasks)
            ]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            has_data = False

            for result in results:
                if isinstance(result, list) and result:
                    has_data = True
                    base_inventario.extend(result)
                elif isinstance(result, Exception):
                    logging.error(f"Error en una de las páginas: {result}")

            if not has_data:
                logging.info("No hay más datos.")
                break

            page += max_concurrent_tasks

    # Eliminar duplicados
    base_inventario = [dict(t) for t in {tuple(d.items()) for d in base_inventario}]
    logging.info(f"Total de registros únicos recuperados: {len(base_inventario)}")
    return base_inventario


def guardar_dataframe(df):
    """
    Guarda el DataFrame en un archivo Excel.
    """
    if os.path.exists(ruta_archivo):
        book = load_workbook(ruta_archivo)
        if hoja_objetivo in book.sheetnames:
            del book[hoja_objetivo]
        book.save(ruta_archivo)
        book.close()

    with pd.ExcelWriter(ruta_archivo, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=hoja_objetivo, index=False)
    logging.info("Datos guardados exitosamente.")


def obtener_dataframe():
    """
    Ejecuta la obtención de datos de la API y los guarda en un archivo Excel.
    """
    datos = asyncio.run(fetch_all_data())

    if not datos:
        logging.warning("No se obtuvieron datos.")
        return pd.DataFrame()

    df = pd.DataFrame(datos)

    # Limpiar y filtrar datos
    df['f120_referencia'] = df['f120_referencia'].str.strip()
    df = df.dropna(subset=['f120_referencia'])

    # Renombrar columnas
    df.columns = [
        'id_company', 'Bodega', 'Item', 'Cod_referencia', 'Ext1', 'Ext2', 'Und',
        'Existencia', 'Existencia2', 'Cantidad_Comprometida', 'Cantidad_comprometida2',
        'Cantidad_salida1', 'Cantidad_salida2', 'Cantidad_pos1', 'Cantidad_post2',
        'Costo_promedio', 'Costo_promedio_total', 'id_lote', 'id_ubicacion'
    ]

    # Guardar el DataFrame
    guardar_dataframe(df)

    return df


if __name__ == "__main__":
    df = obtener_dataframe()
